"""
Create a spreadsheet containing static vs DHCP information for Meraki devices.

Spreadsheet includes index of organizations with hyperlinks to individual org pages.

If MX is acting as the DHCP server for network devices, report shows if a
reservation exists.
"""

import sys
import getopt
import requests
import time
from dataclasses import dataclass
from datetime import datetime
from re import sub
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from string import ascii_uppercase
from netaddr import IPNetwork, IPAddress

# Used for time.sleep(API_EXEC_DELAY). Delay added to avoid hitting
# dashboard API max request rate
API_EXEC_DELAY = 0.21

# connect and read timeouts for the Requests module
REQUESTS_CONNECT_TIMEOUT = 30
REQUESTS_READ_TIMEOUT = 30

# used by merakirequestthrottler(). DO NOT MODIFY
LAST_MERAKI_REQUEST = datetime.now()


@dataclass
class c_orgdata:
    """Class for organization level data."""

    id: str
    name: str
    menu: int
    shard: str


@dataclass
class c_netdata:
    """Class for network level data."""

    id: str
    type: str
    name: str
    tags: str
    menu: int
    orgId: str


@dataclass
class c_devicedata:
    """Class for device level data."""

    serial: str
    name: str
    mac: str
    networkId: str
    model: str
    notes: str
    dhcp: str
    lanIp: str


@dataclass
class c_vlandata:
    """Class for vlan data."""

    id: str
    name: str
    subnet: str
    fixedIpAssignments: dict
    dhcpHandling: str


@dataclass
class c_reservationdata:
    """Class for reservation data."""

    mac: str
    ip: str
    name: str


def printusertext(p_message):
    """Print a line of text meant for the user to read.

    Do not process these lines when chaining scripts
    """
    print(f'@ {p_message}')


def printhelp():
    """Print help text."""
    printusertext('')
    printusertext('merakirequestthrottler(), getorglist(), filterorglist()')
    printusertext('method of getting opts, approach to Meraki API/requests')
    printusertext('use are based off manageadmins.py at')
    printusertext('https://github.com/meraki/automation-scripts/')
    printusertext('by users mpapazog & shiyuechengineer, retrieved 10/19/2018')
    printusertext('')
    printusertext('To run the script, enter:')
    printusertext('python getStaticVsDHCP.py -k <api key> -o <org> -f <file>')
    printusertext('')
    printusertext('-o can be a partial name in quotes with a single wildcard,')
    printusertext('such as \'Calla*\'.')
    printusertext('Use /all for all organizations you have access to.')
    printusertext('')
    printusertext('-f is the file the results will print to.')
    printusertext('')
    printusertext('Use double quotes (/"") in Windows to pass arguments')
    printusertext('containing spaces. Names are case-sensitive.')
    printusertext('')


def merakirequestthrottler(p_requestcount=1):
    """Add delay to requests to avoid hitting shaper."""
    global LAST_MERAKI_REQUEST

    if (datetime.now() - LAST_MERAKI_REQUEST).total_seconds() < (API_EXEC_DELAY * p_requestcount):
        time.sleep(API_EXEC_DELAY * p_requestcount)

    LAST_MERAKI_REQUEST = datetime.now()
    return


def getnetworklist(p_apikey, p_orgid, p_shardurl):
    """Return list of networks for specified organization."""
    merakirequestthrottler()
    r = ""
    try:
        r = requests.get(f"https://{p_shardurl}/api/v0/organizations/{p_orgid}/networks", headers={'X-Cisco-Meraki-API-Key': p_apikey, 'Content-Type': 'application/json'}, timeout=(REQUESTS_CONNECT_TIMEOUT, REQUESTS_READ_TIMEOUT))
    except:
        printusertext('ERROR 02: Unable to contact Meraki cloud')
        sys.exit(2)

    rjson = r.json()

    return(rjson)


def getnetwork(p_apikey, p_networkid, p_shardurl):
    """Get list of networks for an organization."""
    merakirequestthrottler()
    try:
        r = requests.get(f'https://{p_shardurl}/api/v0/networks/{p_networkid}', headers={'X-Cisco-Meraki-API-Key': p_apikey, 'Content-Type': 'application/json'}, timeout=(REQUESTS_CONNECT_TIMEOUT, REQUESTS_READ_TIMEOUT))
    except:
        printusertext('ERROR 02: Unable to contact Meraki cloud')
        sys.exit(2)

    rjson = r.json()

    return(rjson)


def getManagementInterfaceSettings(p_apikey, p_networkid, p_serial, p_shardurl):
    """Get uplink/mgmt IP address."""
    merakirequestthrottler()
    try:
        r = requests.get(f'https://{p_shardurl}/api/v0/networks/{p_networkid}/devices/{p_serial}/managementInterfaceSettings', headers={'X-Cisco-Meraki-API-Key': p_apikey, 'Content-Type': 'application/json'}, timeout=(REQUESTS_CONNECT_TIMEOUT, REQUESTS_READ_TIMEOUT))
    except:
        printusertext('ERROR 02: Unable to contact Meraki cloud')
        sys.exit(2)

    if r:
        rjson = r.json()

    return(rjson)


def getorglist(p_apikey):
    """Return the organizations' list for a specified admin."""
    merakirequestthrottler()
    try:
        r = requests.get('https://api.meraki.com/api/v0/organizations', headers={'X-Cisco-Meraki-API-Key': p_apikey, 'Content-Type': 'application/json'}, timeout=(REQUESTS_CONNECT_TIMEOUT, REQUESTS_READ_TIMEOUT))
    except:
        printusertext('ERROR 01: Unable to contact Meraki cloud')
        sys.exit(2)

    returnvalue = []
    if r.status_code != requests.codes.ok:
        returnvalue.append({'id': 'null'})
        return returnvalue

    rjson = r.json()

    return(rjson)


def getNetworkDevicelist(p_apikey, p_networkid, p_shardurl):
    """Get list of devices for a given network."""
    # returns the Meraki devices in an org
    merakirequestthrottler()
    try:
        r = requests.get(f"https://{p_shardurl}/api/v0/networks/{p_networkid}/devices", headers={'X-Cisco-Meraki-API-Key': p_apikey, 'Content-Type': 'application/json'}, timeout=(REQUESTS_CONNECT_TIMEOUT, REQUESTS_READ_TIMEOUT))
    except:
        printusertext('ERROR 02: Unable to contact Meraki cloud')
        sys.exit(2)

    rjson = r.json()

    return(rjson)


def getNetworkVlans(p_apikey, p_networkid, p_shardurl):
    """Get a list of vlans from a network's MX."""
    merakirequestthrottler()
    try:
        r = requests.get(f"https://{p_shardurl}/api/v0/networks/{p_networkid}/vlans", headers={'X-Cisco-Meraki-API-Key': p_apikey, 'Content-Type': 'application/json'}, timeout=(REQUESTS_CONNECT_TIMEOUT, REQUESTS_READ_TIMEOUT))
    except:
        printusertext('ERROR 02: Unable to contact Meraki cloud')
        sys.exit(2)

    rjson = r.json()

    return(rjson)


def createVlanObj(p_vlanlist):
    """Create a list of vlan objects from JSON API return."""
    returnlist = []
    for vlan in p_vlanlist:
        returnlist.append(c_vlandata(vlan['id'], vlan['name'], vlan['subnet'], vlan['fixedIpAssignments'], vlan['dhcpHandling']))
    return(returnlist)


def createReservationObj(p_reservationlist):
    """Create a list of objects for DHCP reservations from JSON API return."""
    returnlist = []

    for rezKey in p_reservationlist.keys():
        returnlist.append(c_reservationdata(rezKey, '', ''))
        try:
            returnlist[len(returnlist) - 1].name = p_reservationlist[rezKey]['name']
        except:
            pass

        returnlist[len(returnlist) - 1].ip = p_reservationlist[rezKey]['ip']

    if returnlist == []:
        return("No reservations")

    return(returnlist)


def createNetworkObj(p_networklist):
    """Create list of network objects from JSON API return."""
    returnlist = []
    menunum = 1
    for network in p_networklist:
        returnlist.append(c_netdata(network['id'], network['type'], network['name'], network['tags'], menunum, network['organizationId']))
        menunum += 1

    return(returnlist)


def createDeviceObj(p_apikey, p_devicelist, p_shardurl):
    """Create list of device objects based off JSON API return.

    Excludes MV cameras.
    """
    returnlist = []
    for device in p_devicelist:
        if device['model'][0:2] != "MV":
            keys = device.keys()
            uplink = getManagementInterfaceSettings(p_apikey, device['networkId'], device['serial'], p_shardurl)

            # Check if it's a static IP
            if uplink['wan1']['usingStaticIp'] is True:
                returnlist.append(c_devicedata(device['serial'], '', device['mac'], device['networkId'], device['model'], '', 'Static IP', ''))
            else:
                returnlist.append(c_devicedata(device['serial'], '', device['mac'], device['networkId'], device['model'], '', 'DHCP', ''))

            if 'lanIp' in keys:
                returnlist[len(returnlist) - 1].lanIp = device['lanIp']

            if 'name' in keys:
                returnlist[len(returnlist) - 1].name = device['name']
            else:
                returnlist[len(returnlist) - 1].name = "No Name"

            if 'notes' in keys:
                returnlist[len(returnlist) - 1].notes = device['notes']

    # Sort devices by model then SN. Change to suit you.
    sortDevices = sorted(returnlist, key=lambda c_devicedata: (c_devicedata.model[0:2], c_devicedata.serial), reverse=True)
    return(sortDevices)


def checkIpInSubnet(p_subnet, p_ipAddress):
    """Check if device IP is in this subnet."""
    return IPAddress(p_ipAddress) in IPNetwork(p_subnet)


def filterOrgList(p_apikey, p_filter, p_orglist):
    """
    Try to match a list of org IDs to a filter expression.

    Filter:
       /all    all organizations
       <name>  match if name matches. Name can contain one wildcard at start, middle or end.
    """
    returnlist = []

    flag_processall = False
    flag_gotwildcard = False
    if p_filter == '/all':
        flag_processall = True
    else:
        wildcardpos = p_filter.find('*')

        if wildcardpos > -1:
            flag_gotwildcard = True
            startsection = ''
            endsection = ''

            if wildcardpos == 0:
                # Wildcard at start of string, only got endsection.
                endsection = p_filter[1:]

            elif wildcardpos == len(p_filter) - 1:
                # Wildcard at start of string, only got startsection.
                startsection = p_filter[:-1]

            else:
                # Wildcard at middle of string, got both startsection and endsection.
                wildcardsplit = p_filter.split('*')
                startsection = wildcardsplit[0]
                endsection = wildcardsplit[1]

    for org in p_orglist:
        if flag_processall:
            returnlist.append(c_orgdata(org['id'], org['name'], '', org['url']))
        elif flag_gotwildcard:
            flag_gotmatch = True
            # Match startsection and endsection.
            startlen = len(startsection)
            endlen = len(endsection)

            if startlen > 0:
                if org['name'][:startlen] != startsection:
                    flag_gotmatch = False
            if endlen > 0:
                if org['name'][-endlen:] != endsection:
                    flag_gotmatch = False

            if flag_gotmatch:
                returnlist.append(c_orgdata(org['id'], org['name'], '', org['url']))

        else:
            # Match full name.
            if org['name'] == p_filter:
                returnlist.append(c_orgdata(org['id'], org['name'], '', org['url']))

    # Filter org list by name in ascending order.
    sortorgs = []
    sortorgs = sorted(returnlist, key=lambda c_orgdata: c_orgdata.name)
    menunum = 1

    # Set menu number, set shard URL.
    for org in sortorgs:
        org.menu = menunum
        menunum += 1
        urlLength = org.shard.find('com') + 3
        org.shard = org.shard[8:urlLength]
    return(sortorgs)


def main(argv):
    """Create report showing DHCP vs Static IP status on MX, MR, MS."""
    # Initialize variables for command line arguments
    arg_apikey = ''
    arg_orgname = ''
    arg_filename = ''

    # Get command line arguments
    try:
        opts, args = getopt.getopt(argv, 'hk:o:f:')
    except getopt.GetoptError:
        printusertext('Error getting opts.')
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            printhelp()
            sys.exit()
        elif opt == '-k':
            arg_apikey = arg
            if arg_apikey == '':
                printusertext('No API key')
                sys.exit()
        elif opt == '-o':
            arg_orgname = arg
            if arg_orgname == '':
                printusertext('No org name')
                sys.exit()
        elif opt == '-f':
            arg_filename = arg
            if arg_filename == '':
                printusertext('No file name')
                sys.exit()

    # Get list of all orgs this admin has access to.
    raworglist = getorglist(arg_apikey)
    if raworglist[0]['id'] == 'null':
        printusertext('ERROR 08: Error retrieving organization list')
        sys.exit(2)

    # Match list of orgs to org filter
    matchedorgs = filterOrgList(arg_apikey, arg_orgname, raworglist)

    # If no orgs matching - NK 2018-08-23
    if not matchedorgs:
        printusertext(f'Error 10: No organizations matching {arg_orgname}')
        sys.exit(2)

    # Start the workbook
    workbook = Workbook()
    dest_filename = f"{arg_filename}.xlsx"

    # Rename the first sheet
    firstPage = workbook['Sheet']
    firstPage.title = 'OrgList'
    try:
        workbook.save(filename=dest_filename)
    except:
        printusertext(f"Unable to save to {dest_filename}. Check if file is open already.")
        sys.exit(2)

    # Define table style for later...
    style = TableStyleInfo(name="TableStyleLight15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    # Open first sheet. Add headers.
    orgListPage = workbook['OrgList']
    orgListPage.cell(row=1, column=1).value = "List of Organizations On Spreadsheet"

    # Center text
    orgListPage.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    orgListPage.cell(row=1, column=1).font = Font(size="14")
    # Merge cells
    orgListPage.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    orgListPage.cell(row=2, column=1).value = "Meraki Org Name"
    orgListPage.cell(row=2, column=2).value = "Org Tab Name"

    # Initial column width for the index page's A column
    orgListColumnWidth = len("Meraki Org Name")

    # For tracking rows on the first sheet
    orgListNum = 3

    # Header list for all organization sheets
    headerNames = ["Network Name", "Device Name", "Model Type", "Serial Number", "IP Address", "MAC Address", "DHCP Enabled", "Wireless-Only", "DHCP Server", "Reservation"]

    # Set variable for easy reference later
    reservationIndex = headerNames.index("Reservation") + 1

    # Create list of column widths. SN, mac addr, wireless-only and reservation
    # are fixed-width.
    orgSheetColumnWidth = []
    for header in headerNames:
        orgSheetColumnWidth.append(len(header))

    # Iterate through matched orgs
    for org in matchedorgs:
        # Remove non-alphanumeric characters
        cleanOrgName = sub(r'\W+', '', org.name)

        # Create new sheet. Trim name so Excel is happy.
        orgSheet = workbook.create_sheet(f"{cleanOrgName[0:24]}")

        # New worksheet, so set column and row back to start
        columnLetter = 1
        rowNum = 1

        # Add org name to top of sheet, format larger font/center/merged
        orgSheet.cell(row=rowNum, column=columnLetter).value = f"Org Name: {org.name}"
        orgSheet.cell(row=rowNum, column=columnLetter).alignment = Alignment(horizontal="center")
        orgSheet.cell(row=rowNum, column=columnLetter).font = Font(size="14")
        orgSheet.merge_cells(start_row=rowNum, start_column=1, end_row=rowNum, end_column=len(headerNames))
        # Move to next row
        rowNum += 1

        # Add return to table of contents. Advance row.
        orgSheet.cell(row=rowNum, column=columnLetter).value = "=HYPERLINK(\"#OrgList!A1\",\"Return to Org List\")"
        orgSheet.cell(row=rowNum, column=columnLetter).style = "Hyperlink"
        orgSheet.merge_cells(start_row=rowNum, start_column=1, end_row=rowNum, end_column=len(headerNames))
        rowNum += 1

        # Add headers. Advance row.
        for header in headerNames:
            orgSheet.cell(row=rowNum, column=columnLetter).value = header
            columnLetter += 1
        rowNum += 1

        # Make list of network objects
        networkList = createNetworkObj(getnetworklist(arg_apikey, org.id, org.shard))

        print(f"Org Name: {org.name}")
        # Iterate through networks to check devices
        for network in networkList:
            # Set back to start!
            columnLetter = 1
            widthCount = 0

            if len(network.name) > orgSheetColumnWidth[widthCount]:
                orgSheetColumnWidth[widthCount] = len(network.name)
            widthCount += 1
            # Pull list of devices for this network
            deviceList = getNetworkDevicelist(arg_apikey, network.id, org.shard)
            deviceList = createDeviceObj(arg_apikey, deviceList, org.shard)

            # Pull vlan info for this network to check for DHCP reservations
            print(f"\tNetwork name: {network.name}")
            print(f"\tNetwork type: {network.type}")
            print(f"\t\tProcessing devices...\n")
            vlanInfo = getNetworkVlans(arg_apikey, network.id, org.shard)
            try:
                vlanInfo = createVlanObj(vlanInfo)
            except:
                vlanInfo = ""

            # If no applicable devices on this network, tell me that
            if not deviceList:
                print(f"\t\tNo switches or APs found for {network.name}\n")
            else:

                # Populate spreadsheet with device info
                for device in deviceList:
                    # Set width count to element 1, col letter to 1
                    widthCount = 1
                    columnLetter = 1

                    # Set network name
                    orgSheet.cell(row=rowNum, column=columnLetter).value = network.name
                    columnLetter += 1

                    orgSheet.cell(row=rowNum, column=columnLetter).value = device.name
                    # Am I wider? Use widthCount to advance through columns
                    if len(device.name) > orgSheetColumnWidth[widthCount]:
                        orgSheetColumnWidth[widthCount] = len(device.name)

                    # Inc after device name check
                    widthCount += 1
                    columnLetter += 1

                    # Add model and inc columnLetter, widthCount
                    orgSheet.cell(row=rowNum, column=columnLetter).value = device.model
                    # Check width
                    if len(device.model) > orgSheetColumnWidth[widthCount]:
                        orgSheetColumnWidth[widthCount] = len(device.model)
                    columnLetter += 1
                    widthCount += 1

                    # Add the serial number and inc columnLetter, widthCount
                    orgSheet.cell(row=rowNum, column=columnLetter).value = device.serial
                    # Check width
                    if len(device.serial) > orgSheetColumnWidth[widthCount]:
                        orgSheetColumnWidth[widthCount] = len(device.serial)
                    columnLetter += 1
                    widthCount += 1

                    # Add IP address and inc columnLetter, widthCount
                    # AP with no
                    # Fixed width so don't check width
                    if device.lanIp is None:
                        if device.model[0:2] == "MR":
                            orgSheet.cell(row=rowNum, column=columnLetter).value = "Repeater AP?"
                            orgSheet.cell(row=rowNum, column=reservationIndex).value = "Repeater AP?"
                        else:
                            orgSheet.cell(row=rowNum, column=columnLetter).value = "No IP address"
                    else:
                        orgSheet.cell(row=rowNum, column=columnLetter).value = device.lanIp
                    columnLetter += 1
                    widthCount += 1

                    # Add mac and inc columnLetter, widthCount
                    # Fixed width, so don't check width
                    orgSheet.cell(row=rowNum, column=columnLetter).value = device.mac
                    columnLetter += 1
                    widthCount += 1

                    # Am I DHCP or not? Inc columnLetter, widthCount
                    # Fixed width, so don't check width
                    orgSheet.cell(row=rowNum, column=columnLetter).value = device.dhcp
                    columnLetter += 1
                    widthCount += 1

                    # Is this a wireless only network? Inc column, widthCount
                    # Fixed with, so don't check width
                    if network.type == 'wireless':
                        orgSheet.cell(row=rowNum, column=columnLetter).value = 'Yes'
                    else:
                        orgSheet.cell(row=rowNum, column=columnLetter).value = 'No'
                    columnLetter += 1
                    widthCount += 1

                    # Am I given DHCP by this network's MX?
                    # Check if DHCP
                    if device.dhcp != 'DHCP':
                        # Set reservation column to Static IP
                        orgSheet.cell(row=rowNum, column=reservationIndex).value = "Static IP"
                    else:
                        if vlanInfo != "":
                            for vlan in vlanInfo:
                                if vlan.dhcpHandling == "Run a DHCP server":
                                    # Make list of reservations
                                    reservationlist = createReservationObj(vlan.fixedIpAssignments)
                                    if vlan != "":
                                        # Try this because repeaters are on device list but
                                        # don't have an IP address!
                                        try:
                                            if checkIpInSubnet(vlan.subnet, device.lanIp) is True:
                                                # Set DHCP server here since we know MX is running one.
                                                # If there's another DHCP server running, you got problems.
                                                orgSheet.cell(row=rowNum, column=columnLetter).value = f"{network.name} MX"
                                                # Set width.
                                                if len(f"{network.name} MX") > orgSheetColumnWidth[widthCount]:
                                                    orgSheetColumnWidth[widthCount] = len(f"{network.name} MX")

                                                if reservationlist == "No reservations":
                                                    pass
                                                else:
                                                    for reservation in reservationlist:
                                                        # Check if macs match
                                                        if device.mac == reservation.mac:
                                                            orgSheet.cell(row=rowNum, column=reservationIndex).value = reservation.ip
                                                            break
                                        except:
                                            # For this script, do not care if MX is showing as no IP.
                                            pass

                    # Advance rowNum if this row has anything in it.
                    if orgSheet.cell(row=rowNum, column=1).value is not None:
                        rowNum += 1

        # If there aren't any switch or APs, delete the sheet.
        if orgSheet.cell(row=4, column=1).value is None:
            workbook.remove(orgSheet)
            print(f"\t\tNo applicable devices. Deleted {org.name} sheet.")
        else:
            # Write org name to page 1/index. Inc row num.
            orgListPage.cell(row=orgListNum, column=1).value = org.name
            # Adjust column width variable
            if len(org.name) > orgListColumnWidth:
                orgListColumnWidth = len(org.name)

            # Make hyper link so navigating is easier.
            orgListPage.cell(row=orgListNum, column=2).value = f"=HYPERLINK(\"#{cleanOrgName[0:24]}!A1\", \"{cleanOrgName[0:24]}\")"
            orgListPage.cell(row=orgListNum, column=2).style = "Hyperlink"
            orgListNum += 1

            # Set widths on org sheet
            for number in range(len(orgSheetColumnWidth)):
                orgSheet.column_dimensions[ascii_uppercase[number]].width = orgSheetColumnWidth[number] + 5

            # Set table style to white headers, alternating white and light grey rows
            # and black text

            # Use len(headerNames)-1 so you always get a correctly formatted table
            table = Table(displayName=f"{cleanOrgName}DeviceList", ref=f"A3:{ascii_uppercase[len(headerNames)-1]}{rowNum-1}")
            table.tableStyleInfo = style

            # Apply table style to spreadsheet
            orgSheet.add_table(table)

            # Reset column columnWidths
            orgSheetColumnWidth = []
            for header in headerNames:
                orgSheetColumnWidth.append(len(header))
        # Save in case something goes bad
        workbook.save(filename=dest_filename)
        print("Finished processing. Saved workbook.\n\n")

    # Set widths on org list
    orgListPage.column_dimensions['A'].width = orgListColumnWidth * 1.1
    # Fixed width as this can't be longer than 25.
    orgListPage.column_dimensions['B'].width = 28

    # Add table to front page
    orgListTable = Table(displayName="OrgList", ref=f"A2:B{orgListNum-1}")
    orgListTable.tableStyleInfo = style
    # Apply table style to spreadsheet
    orgListPage.add_table(orgListTable)

    # Saving is good
    workbook.save(filename=dest_filename)


if __name__ == '__main__':
    main(sys.argv[1:])
